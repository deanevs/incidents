"""
Plot average KPI per month
No cases per month


"""

import datetime
from openpyxl import load_workbook
from inc_classes import Asset, Incident
from inc_mapping import INC_INT_REF, INC_SAFETY_NOTICE_REF, INC_DATE, REC_DATE,\
     ADDED_TO_AP_DATE, ACTION_DATE, DUE_DATE, CLOSED_DATE, ALERT_FORM, INC_DESC, \
     CLIN_CONSEQ, ACTIONS_TAKEN, DECLARED_BY, ISS_BY, ASSET_ID, EQUIP_NO, GMDN, \
     MANUFACTURE, MODEL, SERIAL
# from inc_mapping import *

# Using the read_only method since you're not gonna be editing the spreadsheet
workbook = load_workbook(filename="Incidents.xlsx", read_only=True)
sheet = workbook.active

incidents_list = []
assets_list = []

def str_to_date(cell_val):
    try:
        converted_date = datetime.datetime.strptime(cell_val,'%Y%m%d')
        return converted_date
    except:
        return None

# Using the values_only because you just want to return the cell value
for row in sheet.iter_rows(min_row=2, values_only=True):
    # print(row[INC_INT_REF])
    incident = Incident(
        int_ref=row[INC_INT_REF],
        safety_not_ref=row[INC_SAFETY_NOTICE_REF],
        inc_date=str_to_date(row[INC_DATE]),
        rec_date=str_to_date(row[REC_DATE]),
        added_date=str_to_date(row[ADDED_TO_AP_DATE]),
        action_date=str_to_date(row[ACTION_DATE]),
        due_date=str_to_date(row[DUE_DATE]),
        closed_date=str_to_date(row[CLOSED_DATE]),
        alert_form=row[ALERT_FORM],
        inc_desc=row[INC_DESC],
        clin_conseq=row[CLIN_CONSEQ],
        actions_taken=row[ACTIONS_TAKEN],
        declared_by=row[DECLARED_BY],
        issued_by=row[ISS_BY]
    )
    incidents_list.append(incident)

    asset = Asset(
        asset_id=row[ASSET_ID],
        equip_no=row[EQUIP_NO],
        gmdn=row[GMDN],
        manu=row[MANUFACTURE],
        model=row[MODEL],
        serial=row[SERIAL]
    )
    assets_list.append(asset)


print(incidents_list[0])
print(assets_list[0])
#print(incidents_list[3].delta_days())

for inc in incidents_list:
    print(inc.delta_days())