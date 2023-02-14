
import json
from openpyxl import load_workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import PatternFill, colors

workbook = load_workbook(filename="Incidents_1.xlsx")
sheet = workbook.active

cnt = 0
for row in sheet.iter_rows():
    cnt = cnt + 1
    cell_kpi = "I{}".format(cnt)
    cell_due_days = "L{}".format(cnt)

    elapsed_workdays_add_ap = row[8].value
    ewa = None
    color = colors.RED

    try:
        ewa = int(elapsed_workdays_add_ap)
        if ewa > 5:
            color = colors.RED
            end_val = ewa
        else:
            color = colors.GREEN
            end_val = 5
    except:
        pass

    data_bar_rule = DataBarRule(start_type="num",
                                start_value=1,
                                end_type="num",
                                end_value= 5,
                                color=color
                                )
    try:
        print(cell_kpi)
        if isinstance(end_val, int):
            sheet.conditional_formatting.add(cell_kpi, data_bar_rule)
    except:
        pass



# sheet.conditional_formatting.add("H2:H100", data_bar_rule)
workbook.save("incidents_1_conditional_formatting_data_bar_2.xlsx")