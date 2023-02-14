import random
import datetime
from openpyxl import Workbook
import sys
import numpy as np
from inc_filler_mapping import *


def date_to_str(in_date):
    if isinstance(in_date, datetime.date):
        result = in_date.strftime('%Y%m%d')
        return result
    else:
        return ''

def str_to_date(in_date_str):
    result = datetime.strptime(in_date_str,"%Y%m%d").date()
    return result

def calc_elapsed_workdays(start_date, end_date):
    try:
        days_elapsed = np.busday_count(start_date, end_date)
        print(days_elapsed)
        return days_elapsed
    except Exception as e:
        print(e)
    # finally:
    #     days_elapsed = 'error'
    #     return days_elapsed


inc_list = []   # holds all the new incident dates so that we can sort

# init all the date types we want to fill
inc_date = datetime    # any time in the last 365 days
rec_date = datetime    # inc_date + (0 - 5)
add_date = datetime    # rec_date + (0 - 8)
due_date = datetime    # inc_date + ((4 - 8) * 7)  ie set no weeks
clo_date = datetime    # inc_date + (20 - 60)

# generate list of random dates back from today
DAYS_BEHIND = 365
NO_ROWS = 100

for x in range(NO_ROWS):
    incident_date = datetime.date.today() - datetime.timedelta(days=random.randint(0,DAYS_BEHIND))
    inc_list.append(incident_date)

# sort the dates
inc_list.sort(reverse=True)

# holder for a list of lists for all dates for each row
all_dates = []

# create dates according to formulae in init sections for each row
for inc in inc_list:
    temp = []   # init temporary list to hold all dates for each row
    temp.append(inc)
    rec_date = inc + datetime.timedelta(days=random.randint(0,5))
    temp.append(rec_date)
    add_date = rec_date + datetime.timedelta(random.randint(0,8))
    temp.append(add_date)
    due_date = inc + datetime.timedelta(random.randint(4,8) * 7)
    temp.append(due_date)
    # closed cannot be in the future
    clo_date = inc + datetime.timedelta(random.randint(20,60))
    if clo_date <= datetime.date.today():
        temp.append(clo_date)
    else:
        temp.append("")

    all_dates.append(temp)

# Set up excel workbook

wb = Workbook()
ws = wb.active

# fill headers
ws.append(['Internal Ref', 'Status', 'Safety Notice Ref', 'AssetPlus Identifier', 'Equipment No', 'Incident Date',
           'Received Date', 'Added to Assetplus Date', 'Elapsed Workdays', 'Action Date', 'Due Date',
           'Workdays to Due Date', 'Overdue Status', 'Closed Date', 'Performance', 'Alert Form', 'Incident Description',
           'Clinical Consequences', 'Actions Taken', 'GMDN', 'Manufacturer', 'Model', 'Serial', 'Declared By',
           'Issued By'])

# now fill the sheet
start_row = 2   # headers row 1!
end_row = NO_ROWS + 1   # account for header

today = datetime.date.today()

row_cnt = 0
for row in range(start_row, end_row + 1, 1):
    ws.cell(row=row, column=INT_REF, value="INT REF {}".format(row))
    ws.cell(row=row, column=STATUS, value="STATUS {}".format(row))
    ws.cell(row=row, column=SAFETY_REF, value="SAFETY REF {}".format(row))
    ws.cell(row=row, column=ASSET_ID, value="50000{}".format(row))
    ws.cell(row=row, column=EQUIP_NO, value="50000{}".format(row))
    ws.cell(row=row, column=INT_REF, value="INT REF {}".format(row))
    ws.cell(row=row, column=INC_DATE, value=date_to_str(all_dates[row_cnt][0]))
    ws.cell(row=row, column=REC_DATE, value=date_to_str(all_dates[row_cnt][1]))
    ws.cell(row=row, column=ADDED_DATE, value=date_to_str(all_dates[row_cnt][2]))
    # this will always be finalised since the rec and add will be done the same time
    ws.cell(row=row, column=ELAPSED_DAYS, value=calc_elapsed_workdays(all_dates[row_cnt][1],all_dates[row_cnt][2]))
    ws.cell(row=row, column=ACT_DATE, value="")
    ws.cell(row=row, column=DUE_DATE, value=date_to_str(all_dates[row_cnt][3]))
    # this will be from todays date
    if isinstance(all_dates[row_cnt][4], datetime.date): # if already closed
        ws.cell(row=row, column=WORKDAYS_DUE, value="")
        ws.cell(row=row, column=CLOSED_DATE, value=date_to_str(all_dates[row_cnt][4]))
        ws.cell(row=row, column=PERFORMANCE, value=calc_elapsed_workdays(all_dates[row_cnt][3], all_dates[row_cnt][4]))
    else:
        ws.cell(row=row, column=WORKDAYS_DUE, value=calc_elapsed_workdays(today, all_dates[row_cnt][3]))
        ws.cell(row=row, column=PERFORMANCE, value="")

    # now increment the row counter
    row_cnt += 1

wb.save(filename="incident_test_1.xlsx")

